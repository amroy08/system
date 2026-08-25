import argparse
import json
from pathlib import Path

from openpyxl import load_workbook


REQUIRED_COLUMNS = {
    "SN",
    "Name",
    "Class",
    "Contact",
    "Fees",
    "Old Balance",
    "Total",
    "Received",
    "Outstsnding",
    "Last paid",
}


def extract_workbook(source: Path) -> list[dict]:
    workbook = load_workbook(source, read_only=True, data_only=True)
    worksheet = workbook.active
    rows = worksheet.iter_rows(values_only=True)
    next(rows, None)  # Export title row
    headers = [str(value).strip() if value is not None else "" for value in next(rows, ())]
    missing = REQUIRED_COLUMNS.difference(headers)
    if missing:
        workbook.close()
        raise ValueError(f"{source.name}: missing columns: {', '.join(sorted(missing))}")

    records = []
    for source_row, values in enumerate(rows, start=3):
        record = dict(zip(headers, values))
        if not record.get("Name"):
            continue
        record["_sourceRow"] = source_row
        records.append(record)

    workbook.close()
    return records


def main() -> None:
    parser = argparse.ArgumentParser(description="Extract validated MVHS student workbooks to JSON.")
    parser.add_argument("--primary", type=Path, required=True)
    parser.add_argument("--secondary", type=Path, required=True)
    parser.add_argument("--old", type=Path, required=True)
    parser.add_argument("--output", type=Path, required=True)
    args = parser.parse_args()

    args.output.mkdir(parents=True, exist_ok=True)
    for label in ("primary", "secondary", "old"):
        source = getattr(args, label)
        records = extract_workbook(source)
        destination = args.output / f"{label}.json"
        destination.write_text(json.dumps(records, indent=2, ensure_ascii=False), encoding="utf-8")
        print(f"{label}: {len(records)} records")


if __name__ == "__main__":
    main()
